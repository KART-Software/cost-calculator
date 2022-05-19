from cost_calculator.fca import FcaSheet
from typing import List
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from os.path import relpath

from cost_calculator import FcaSheet
from cost_calculator.categories import CostCategory, SystemAssemblyCategory
import openpyxl


class BomSheet:
    filePath: str
    bomBook: Workbook
    bomSheet: Worksheet
    isNotBomSheet: bool
    baseRow: int
    costColumns: List[int]
    asmPrtColumn: int
    componentColumn: int
    quantityColumn: int
    linkToFcaSheetColumn: int
    systemAssemblyRowRanges: List[tuple]

    def __init__(self, path: str):
        self.filePath = path
        self.bomBook = openpyxl.load_workbook(path)
        self.bomSheet = self.bomBook.worksheets[1]
        self._detectBaseRowAndColumns()
        if self.isNotBomSheet == False:
            self._detectSystemAssemblyRowRanges()

    def _detectBaseRowAndColumns(self):
        self.isNotBomSheet = True
        for row in range(1, 15):
            if self.bomSheet.cell(row, 1).value == 1:
                if self.bomSheet.cell(row + 1, 1).value == 2:
                    self.baseRow = row - 1
                    self.isNotBomSheet = False
                    break
        if self.isNotBomSheet == False:
            self.costColumns = [None, None, None, None, None]
            for column in range(1, self.bomSheet.max_column + 1):
                cellValue = self.bomSheet.cell(self.baseRow, column).value
                if cellValue == "Asm/Prt #":
                    self.asmPrtColumn = column
                if cellValue == "Component":
                    self.componentColumn = column
                if cellValue == "Quantity":
                    self.quantityColumn = column
                if cellValue == CostCategory.Material.categoryName + " Cost":
                    self.costColumns[CostCategory.Material] = column
                    self.costColumns[CostCategory.Process] = column + 1
                    self.costColumns[CostCategory.Fastener] = column + 2
                    self.costColumns[CostCategory.Tooling] = column + 3
                if cellValue == "Link to FCA Sheet":
                    self.linkToFcaSheetColumn = column

    def _detectSystemAssemblyRowRanges(self):
        self.systemAssemblyRowRanges = [
            None, None, None, None, None, None, None, None
        ]
        startRow = self.baseRow + 1
        for row in range(startRow, self.bomSheet.max_row + 1):
            if self.bomSheet.cell(row, 1).value == None:
                for category in SystemAssemblyCategory:
                    if self.bomSheet.cell(row,
                                          2).value in category.categoryName:
                        endRow = row - 1
                        self.systemAssemblyRowRanges[category] = (startRow,
                                                                  endRow)
                        startRow = row + 1
                        break

    def enterData(self, fcaSheet: FcaSheet) -> bool:
        rowRange = self.systemAssemblyRowRanges[
            fcaSheet.systemAssemblyCategory]
        component = fcaSheet.title
        asmPrt = fcaSheet.id
        linkToFcaSheet = relpath(fcaSheet.fcaFilePath, self.filePath + "/..")
        entered = False

        for row in range(rowRange[0], rowRange[1] + 1):
            if str(self.bomSheet.cell(
                    row, self.asmPrtColumn).value).lower() == asmPrt.lower():
                # Quantity
                if fcaSheet.isAsmSheet:
                    quantity = fcaSheet.getQuantity()
                    if quantity == None:
                        print("FcaファイルにQtyが入っていません, ", end="")
                    else:
                        print("Quantity OK!!, ", end="")
                else:
                    asmId = "A" + fcaSheet.id[:3] + "0"
                    if fcaSheet.asmQuantities[asmId] != None:
                        quantity_ = fcaSheet.getQuantity()
                        if quantity_ == None:
                            print("FcaファイルにQtyが入っていません, ", end="")
                        else:
                            print("Quantity OK!!, ", end="")
                        quantity = quantity_ * fcaSheet.asmQuantities[asmId]
                    else:
                        print("Quantityを決定できません,  ", end="")
                        quantity = None
                self.bomSheet.cell(row, self.quantityColumn, value=quantity)
                # Costs
                for category in CostCategory:
                    if category != CostCategory.ProcessMultiplier:
                        self.bomSheet.cell(
                            row,
                            self.costColumns[category],
                            value=fcaSheet.getSubTotal(category))
                print("Costs OK!!, ", end="")
                # Link to Fca Sheet
                linkName = str(
                    self.bomSheet.cell(row, self.asmPrtColumn).value)
                hyperLink = "=HYPERLINK(\"[{}]\'{}\'!A1\",\"{}\")".format(
                    linkToFcaSheet, asmPrt, linkName)
                # hyperLink = "=HYPERLINK(\"[" + linkToFcaSheet + "]\'" + component + "\'!A1\",\"" + linkName + "\")"
                self.bomSheet.cell(row,
                                   self.linkToFcaSheetColumn,
                                   value=hyperLink)
                print("Link to FCA OK!! : ", end="")
                entered = True
        if entered == False:
            print("FCAのシート名と一致するBOMの行を見つけられませんでした。 : ", end="")
        print("{} : {} :".format(fcaSheet.id, fcaSheet.title))
        return entered

    # def deleteData(self):
    #     for category in SystemAssemblyCategory:

    def save(self):
        self.bomBook.save(self.filePath)