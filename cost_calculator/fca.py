from glob import glob
from os.path import relpath
from typing import List
import openpyxl
from openpyxl.workbook.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from cost_calculator import CostTable
from cost_calculator.categories import Cost, CostCategory, SystemAssemblyCategory
from cost_calculator.supplement import SupplPdf


def supplToFca(fcaDirectoryPath):

    fcaFilePaths = glob(fcaDirectoryPath + "/*.xlsx")
    fcaFilePaths.extend(glob(fcaDirectoryPath + "/*/*.xlsx"))
    fcaFilePaths.extend(glob(fcaDirectoryPath + "/*/*/*.xlsx"))
    for path in fcaFilePaths:
        fca = Fca(path, parseSupplPdf=True)
        if fca.isFca:
            if fca.hasSupplPdf:
                fca.enterLinkToSuppleent()
                fca.save()


class FcaSheet:
    CATEGORY_COLUMN = 2
    UNIT_COST_COLUMN = 4
    MULTIPLIER_COLUMN = 7
    MULTVAL_COLUMN = 8
    CATEGORY_ROW_TO_CHECK_FROM = 9
    ID_COLUMN = 2
    SYSTEM_ASSEMBLY_CATEGORY_CELL = (2, 2)
    QUANTITY_CELL = (2, 14)
    FILE_LINK_CELL = (3, 11)
    #TODO これらのセル、行、列を自動で検知するように
    #TODO idRow、title、isNotFcaSheet、idなどを @propertyに

    fcaSheet: Worksheet
    categoryRowRanges: List[tuple]
    subTotalColumns: List[int]
    idRow: int
    systemAssemblyCategory: SystemAssemblyCategory
    isNotFcaSheet: bool
    fcaFilePath: str

    def __init__(self, fcaSheet: Worksheet):
        self.fcaSheet = fcaSheet
        self._detectSystemAssemblyCategory()
        if self.isNotFcaSheet == False:
            self._detectCategoryRowRanges()
            self._detectSubTotalColumns()
            self._detectIdRow()

    def _detectSystemAssemblyCategory(self):
        self.isNotFcaSheet = True
        cellValue = self.fcaSheet.cell(
            FcaSheet.SYSTEM_ASSEMBLY_CATEGORY_CELL[0],
            FcaSheet.SYSTEM_ASSEMBLY_CATEGORY_CELL[1]).value
        for category in SystemAssemblyCategory:
            if cellValue in category.categoryName:
                self.systemAssemblyCategory = category
                self.isNotFcaSheet = False

    def _detectCategoryRowRanges(self):
        category: CostCategory
        startRow: int
        self.categoryRowRanges = [None, None, None, None, None]
        startRow = None
        for row in range(FcaSheet.CATEGORY_ROW_TO_CHECK_FROM,
                         self.fcaSheet.max_row + 1):
            cellValue = self.fcaSheet.cell(row, FcaSheet.CATEGORY_COLUMN).value
            if cellValue in Cost.CATEGORY_NAMES:
                category = Cost.CATEGORY_NAMES.index(cellValue)
                startRow = row
            if startRow:
                if row > startRow and cellValue == None:
                    endRow = row - 1
                    self.categoryRowRanges[category] = (startRow, endRow)
                    startRow = None

    def _detectSubTotalColumns(self):
        self.subTotalColumns = list(range(5))
        self.subTotalColumns[CostCategory.ProcessMultiplier] = None
        for category in CostCategory:
            if category != CostCategory.ProcessMultiplier:
                for column in range(1, self.fcaSheet.max_column + 1):
                    if self.fcaSheet.cell(self.categoryRowRanges[category][0],
                                          column).value == "Sub Total":
                        self.subTotalColumns[category] = column
                        column += 1
                        break
                    column += 1

    def _detectIdRow(self):
        for row in range(1, 10):
            if self.fcaSheet.cell(row,
                                  FcaSheet.ID_COLUMN - 1).value == "P/N Base":
                self.idRow = row

    def putFcaData(self, fcaFilePath, supplPdf: SupplPdf, hasSupplPdf: bool):
        self.fcaFilePath = fcaFilePath
        self.supplPdf = supplPdf
        self.hasSupplPdf = hasSupplPdf

    def enterCost(self, category: CostCategory, costTable: CostTable):
        if category == CostCategory.Process:
            # error
            pass
        #TODO forを使って書き直し
        row = self.categoryRowRanges[category][0] + 1
        while True:
            if (self.fcaSheet.cell(row,
                                   FcaSheet.CATEGORY_COLUMN).value == None):
                break
            if (self.fcaSheet.cell(row,
                                   FcaSheet.CATEGORY_COLUMN).value == "None"):
                self.fcaSheet.cell(row, FcaSheet.UNIT_COST_COLUMN, value=0)
                break
            self.fcaSheet.cell(row,
                               FcaSheet.UNIT_COST_COLUMN,
                               value=costTable.getCost(
                                   self.fcaSheet.cell(
                                       row, FcaSheet.CATEGORY_COLUMN).value))
            row += 1

    def enterProcessCost(self, tableProcesses: CostTable,
                         tableProcessMultipliers: CostTable):
        MULTIPLIER_PREFIXES = ["", "Machine - ", "Material - "]
        #TODO forを使って書き直し
        row = self.categoryRowRanges[CostCategory.Process][0] + 1
        while True:
            if (self.fcaSheet.cell(row,
                                   FcaSheet.CATEGORY_COLUMN).value == None):
                break
            cost = tableProcesses.getCost(
                self.fcaSheet.cell(row, FcaSheet.CATEGORY_COLUMN).value)
            self.fcaSheet.cell(row, FcaSheet.UNIT_COST_COLUMN, value=cost)
            if self.fcaSheet.cell(row,
                                  FcaSheet.MULTIPLIER_COLUMN).value == None:
                multiplier = Cost(1.0)
            else:
                multiplier = tableProcessMultipliers.getCost(
                    self.fcaSheet.cell(row, FcaSheet.MULTIPLIER_COLUMN).value)
                for prefix in MULTIPLIER_PREFIXES:
                    multiplier_ = tableProcessMultipliers.getCost(
                        prefix + self.fcaSheet.cell(
                            row, FcaSheet.MULTIPLIER_COLUMN).value)
                    if multiplier_ != None:
                        multiplier = multiplier_
                        break
                self.fcaSheet.cell(row,
                                   FcaSheet.MULTVAL_COLUMN,
                                   value=multiplier)
            row += 1

    def deleteCost(self, category: CostCategory):
        if category == CostCategory.Process:
            # error
            pass
        row = self.categoryRowRanges[category][0] + 1
        #TODO forを使って書き直し
        while True:
            if (self.fcaSheet.cell(row,
                                   FcaSheet.CATEGORY_COLUMN).value == None):
                break
            self.fcaSheet.cell(row, FcaSheet.UNIT_COST_COLUMN, value="")
            row += 1

    def deleteProcessCost(self):
        row = self.categoryRowRanges[CostCategory.Process][0] + 1
        #TODO forを使って書き直し
        while True:
            if (self.fcaSheet.cell(row,
                                   FcaSheet.CATEGORY_COLUMN).value == None):
                break
            self.fcaSheet.cell(row, FcaSheet.UNIT_COST_COLUMN, value="")
            self.fcaSheet.cell(row, FcaSheet.MULTVAL_COLUMN, value="")
            row += 1

    def enterLinkToSupplement(self):
        # id = str(self.fcaSheet.cell(self.idRow, FcaSheet.ID_COLUMN).value)
        id = self.id
        directoryPath = relpath(self.fcaFilePath + "/..")
        if self.hasSupplPdf:
            linkToPdf = relpath(self.supplPdf.filePath, directoryPath)
            page = self.supplPdf.pageOfId(id)
            if page:
                hyperLink = "=HYPERLINK(\"{}#page={}\")".format(
                    linkToPdf, page)
                # hyperLink = "=HYPERLINK(\"{}#page={}\",\"{}\")".format(
                #     linkToPdf, page, id)
                self.fcaSheet.cell(FcaSheet.FILE_LINK_CELL[0],
                                   FcaSheet.FILE_LINK_CELL[1] + 1,
                                   value="Page=" + str(page))
                print("OK!! : {} : {}".format(id, self.fcaSheet.title))
            else:
                hyperLink = "=HYPERLINK(\"{}\")".format(linkToPdf)
                # hyperLink = "=HYPERLINK(\"{}\",\"{}\")".format(linkToPdf, id)
                print("裏付け資料に、IDが一致するページがありません。 : {} : {}".format(
                    id, self.fcaSheet.title))
            self.fcaSheet.cell(FcaSheet.FILE_LINK_CELL[0],
                               FcaSheet.FILE_LINK_CELL[1],
                               value=hyperLink)

    def getQuantity(self):
        return self.fcaSheet.cell(self.QUANTITY_CELL[0],
                                  self.QUANTITY_CELL[1]).value

    def getSubTotal(self, category: CostCategory) -> Cost:
        cellvalue = self.fcaSheet.cell(self.categoryRowRanges[category][1] + 1,
                                       self.subTotalColumns[category]).value
        if type(cellvalue) != int and type(cellvalue) != float:
            return None
        return Cost(float(cellvalue))

    @property
    def title(self) -> str:
        return self.fcaSheet.title

    @property
    def id(self) -> str:
        return str(self.fcaSheet.cell(self.idRow, FcaSheet.ID_COLUMN).value)


class Fca:
    MUST_INCLUDE_CELL = {
        (1, 1): "University",
        (2, 1): "System",
        (3, 1): "Assembly"
    }
    isFca: bool
    fcaSheets: List[FcaSheet]
    filePath: str
    fcaBook: Workbook
    supplePdf: SupplPdf
    hasSupplPdf: bool

    def __init__(self, path: str, data_only=False, parseSupplPdf=False):
        self.filePath = path
        self.supplePdf = None
        self.hasSupplPdf = False
        self.fcaBook = openpyxl.load_workbook(path, data_only=data_only)
        self._judge()
        if self.isFca:
            if parseSupplPdf:
                self._parseSupplPdf()
            self.fcaSheets = []
            for sheet in self.fcaBook.worksheets:
                fcaSheet = FcaSheet(sheet)
                if fcaSheet.isNotFcaSheet == False:
                    fcaSheet.putFcaData(self.filePath,
                                        supplPdf=self.supplePdf,
                                        hasSupplPdf=self.hasSupplPdf)
                    self.fcaSheets.append(fcaSheet)

    def _judge(self):
        self.isFca = True
        sheet0 = self.fcaBook.worksheets[0]
        for cell, value in Fca.MUST_INCLUDE_CELL.items():
            self.isFca = self.isFca and sheet0.cell(cell[0],
                                                    cell[1]).value == value

    def _parseSupplPdf(self):
        directoryPath = relpath(self.filePath + "/..")
        pdfPaths = glob(directoryPath + "/*.pdf")
        supplePdfs = []
        for path in pdfPaths:
            pdf = SupplPdf(path)
            if pdf.isSupplPDF:
                supplePdfs.append(pdf)
        if len(supplePdfs) == 1:
            self.supplePdf = supplePdfs[0]
            self.hasSupplPdf = True
        else:
            print("次のFCAファイルの裏付け資料が複数存在、または存在しません。\
                    FCAファイルと同じディレクトリに１つだけ置いてください。")
            print(self.filePath)
            self.supplePdf = None
            self.hasSupplPdf = False

    def enterLinkToSuppleent(self):
        for sheet in self.fcaSheets:
            sheet.enterLinkToSupplement()

    def save(self):
        self.fcaBook.save(self.filePath)