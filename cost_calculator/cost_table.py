from enum import IntEnum
import openpyxl
from openpyxl.worksheet.worksheet import Worksheet
from cost_calculator.categories import Cost, CostCategory


class CostTable:
    # GENERICTERM_VALUENAME_SHEETTITLE = {
    #     CostCategory.Material:
    #     ("Material", ("Table Price", "Calc Value"), "tblMaterials"),
    #     CostCategory.Process: ("Process", ("Unit Cost", ), "tblProcesses"),
    #     CostCategory.ProcessMultiplier:
    #     ("Process Multiplier", ("Multiplier", ), "tblProcessMultipliers"),
    #     CostCategory.Fastener:
    #     ("Fastener", ("Table Price", "Calc Price"), "tblFasteners"),
    #     CostCategory.Tooling: ("Process", ("Cost", ), "tblToolings"),
    # }
    GENERIC_TERM = [
        "Material", "Process", "Process Multiplier", "Fastener", "Process"
    ]
    VALUE_NAME = [("Table Price", "Calc Value"), ("Unit Cost", ),
                  ("Multiplier", ), ("Table Price", "Calc Price"), ("Cost", )]
    SHEET_TITLE = [
        "tblMaterials", "tblProcesses", "tblProcessMultipliers",
        "tblFasteners", "tblTooling"
    ]

    GENERIC_TERM_COLUMN = 2

    costSheet: Worksheet
    isNotCostTable: bool
    category: CostCategory
    baseRow: int
    valueCollums: tuple

    def __init__(self, path: str):
        self.costSheet = openpyxl.load_workbook(path,
                                                data_only=True).worksheets[0]
        self._detectCategory()
        if self.isNotCostTable == False:
            self._detectBaseRowAndCollum()

    def _detectCategory(self):
        self.isNotCostTable = True
        for category in CostCategory:
            if self.costSheet.title == CostTable.SHEET_TITLE[category]:
                self.category = category
                self.isNotCostTable = False
                break
            self.isNotCostTable = self.isNotCostTable and self.costSheet.title != CostTable.SHEET_TITLE[
                category]

    def _detectBaseRowAndCollum(self):
        for i in range(1, 5):
            if (self.costSheet.cell(i, CostTable.GENERIC_TERM_COLUMN).value ==
                    CostTable.GENERIC_TERM[self.category]):
                self.baseRow = i
                break
            if i >= 4:
                # error
                break
        columns = []
        for column in range(1, self.costSheet.max_column):
            if self.costSheet.cell(
                    self.baseRow,
                    column).value in CostTable.VALUE_NAME[self.category]:
                columns.append(column)
        self.valueCollums = tuple(columns)

    def getCost(self, costName: str) -> Cost:
        for row in range(self.baseRow + 1, self.costSheet.max_row + 1):
            if self.costSheet.cell(
                    row, CostTable.GENERIC_TERM_COLUMN).value == None:
                # error
                break
            if self.costSheet.cell(
                    row, CostTable.GENERIC_TERM_COLUMN).value == costName:
                for column in self.valueCollums:
                    if (type(self.costSheet.cell(row, column).value) == float
                            or type(self.costSheet.cell(row,
                                                        column).value) == int):
                        return Cost(
                            float(self.costSheet.cell(row, column).value))